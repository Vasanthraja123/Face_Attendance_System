import os
import base64
import numpy as np
import cv2
from datetime import datetime
from flask import Flask, request, jsonify, render_template
from tensorflow.keras.applications import MobileNetV2
from tensorflow.keras.applications.mobilenet_v2 import preprocess_input
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.image import img_to_array
from tensorflow.keras.layers import Dense, GlobalAveragePooling2D
from tensorflow.keras.models import Model
from tensorflow.keras.preprocessing import image
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Configuration
UPLOAD_FOLDER = 'Dataset'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

ENTRY_LOG_FILE = "entry_log.xlsx"
EXIT_LOG_FILE = "exit_log.xlsx"
MOBILENET_MODEL_PATH = "mobilenetv2_model.h5"

# Class mapping
KNOWN_FACES = {
    "classes": [
        "Sakthivel",
        "Siddharth",
        "Vasanth Kumar"
    ]
}

class FaceRecognition:
    def __init__(self, model_path):
        print("Loading MobileNetV2 model from:", model_path)
        base_model = MobileNetV2(weights=None, include_top=False, input_shape=(224, 224, 3))
        x = GlobalAveragePooling2D()(base_model.output)
        x = Dense(len(KNOWN_FACES["classes"]), activation='softmax')(x)
        self.model = Model(inputs=base_model.input, outputs=x)
        self.model.load_weights(model_path)
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        
        print("Model output shape:", self.model.output_shape)
        print("Number of classes:", len(KNOWN_FACES["classes"]))

    def process_image(self, image_data):
        try:
            img_data = base64.b64decode(image_data.split(",")[1])
            nparr = np.frombuffer(img_data, np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            if frame is None:
                raise ValueError("Failed to decode image")
            return frame
        except Exception as e:
            raise ValueError(f"Error processing image: {str(e)}")

    def detect_faces(self, frame):
        try:
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5, minSize=(50, 50))
            print(f"Detected {len(faces)} faces")
            return faces
        except Exception as e:
            print(f"Error detecting faces: {str(e)}")
            return []

    def predict_face(self, face_crop):
        try:
            face_resized = cv2.resize(face_crop, (224, 224))
            face_array = img_to_array(face_resized)
            face_array = preprocess_input(face_array)
            face_array = np.expand_dims(face_array, axis=0)

            predictions = self.model.predict(face_array, verbose=0)
            class_index = np.argmax(predictions[0])
            confidence = predictions[0][class_index]
            
            return class_index, confidence
        except Exception as e:
            print(f"Error in predict_face: {str(e)}")
            return None, 0.0

    def recognize_face(self, image_data):
        print("\nStarting face recognition process...")
        frame = self.process_image(image_data)
        faces = self.detect_faces(frame)
        
        results = []
        if len(faces) == 0:
            print("No faces detected in the image")
            return [("No face detected", None, 0)], frame

        for (x, y, w, h) in faces:
            face_crop = frame[y:y+h, x:x+w]
            class_index, confidence = self.predict_face(face_crop)
            
            if class_index is not None and confidence > 0.5:
                name = KNOWN_FACES["classes"][class_index]
                results.append((name, (x, y, w, h), confidence))
            else:
                results.append(("Unknown", (x, y, w, h), confidence))

        return results, frame

class AttendanceLog:
    def __init__(self, entry_log_file, exit_log_file):
        self.entry_log_file = entry_log_file
        self.exit_log_file = exit_log_file
        self._initialize_log_files()

    def _initialize_log_files(self):
        for log_file in [self.entry_log_file, self.exit_log_file]:
            if not os.path.exists(log_file):
                wb = Workbook()
                ws = wb.active
                ws.append(["Name", "Date", "Timestamp"])
                wb.save(log_file)

    def log_attendance(self, name, action):
        log_file = self.entry_log_file if action == "enter" else self.exit_log_file
        now = datetime.now()
        date_today = now.strftime("%Y-%m-%d")
        time_stamp = now.strftime("%Y-%m-%d %H:%M:%S")
        
        try:
            wb = load_workbook(log_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == name and row[1] == date_today:
                    return False

            ws.append([name, date_today, time_stamp])
            wb.save(log_file)
            return True
        except Exception as e:
            print(f"Error logging attendance: {e}")
            return False

class FaceRecognitionApp:
    def __init__(self, app, model_path, entry_log_file, exit_log_file):
        self.app = app
        self.face_recognition = FaceRecognition(model_path)
        self.attendance_log = AttendanceLog(entry_log_file, exit_log_file)
        self.setup_routes()

    def setup_routes(self):
        self.app.route("/")(self.index)
        self.app.route("/recognize_face", methods=["POST"])(self.recognize_face)

    def index(self):
        return render_template("index.html")

    def recognize_face(self):
        try:
            data = request.json
            image_data = data.get("image")
            action = data.get("action")
            if not image_data or not action:
                return jsonify({"error": "Invalid data"}), 400
            
            results, frame = self.face_recognition.recognize_face(image_data)
            response_data = [{"name": name, "confidence": float(confidence), "status": "logged" if self.attendance_log.log_attendance(name, action) else "already logged"} for name, _, confidence in results if name != "Unknown"]
            return jsonify({"message": "Attendance processed", "results": response_data}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500

face_recognition_app = FaceRecognitionApp(app, MOBILENET_MODEL_PATH, ENTRY_LOG_FILE, EXIT_LOG_FILE)

if __name__ == "__main__":
    app.run(debug=True)
