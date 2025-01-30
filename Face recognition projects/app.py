import os
import base64
import numpy as np
import cv2
from datetime import datetime
from flask import Flask, request, jsonify, render_template
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.image import img_to_array
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Configuration
UPLOAD_FOLDER = 'Dataset'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

ENTRY_LOG_FILE = "entry_log.xlsx"
EXIT_LOG_FILE = "exit_log.xlsx"
CNN_MODEL_PATH = "cnn_model_4_classes (crop).h5"

# Class mapping
KNOWN_FACES = {
    "classes": [
        "Sakthivel",
        "Siddharth",
        "Vasanth Kumar"
    ]
}

class FaceRecognition:
    def __init__(self, model_path):
        print("Loading model from:", model_path)
        self.cnn_model = load_model(model_path)
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        
        # Print model summary and output shape
        self.cnn_model.summary()
        print("Model output shape:", self.cnn_model.output_shape)
        print("Number of classes:", len(KNOWN_FACES["classes"]))

    def process_image(self, image_data):
        try:
            img_data = base64.b64decode(image_data.split(",")[1])
            nparr = np.frombuffer(img_data, np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            if frame is None:
                raise ValueError("Failed to decode image")
            return frame
        except Exception as e:
            raise ValueError(f"Error processing image: {str(e)}")

    def detect_faces(self, frame):
        try:
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(
                gray_frame,
                scaleFactor=1.1,
                minNeighbors=5,
                minSize=(50, 50)
            )
            print(f"Detected {len(faces)} faces")
            return faces
        except Exception as e:
            print(f"Error detecting faces: {str(e)}")
            return []

    def predict_face(self, face_crop):
        try:
            # Preprocess the face image
            input_size = self.cnn_model.input_shape[1:3]
            face_resized = cv2.resize(face_crop, input_size)
            face_array = img_to_array(face_resized)
            face_array = face_array / 255.0  # Normalize
            face_array = np.expand_dims(face_array, axis=0)

            # Get predictions
            predictions = self.cnn_model.predict(face_array, verbose=0)
            print("Raw predictions shape:", predictions.shape)
            print("Raw predictions:", predictions)

            # Get class index and confidence
            class_index = np.argmax(predictions[0])
            confidence = predictions[0][class_index]

            # Print detailed prediction information
            print(f"Predicted class index: {class_index}")
            print(f"Confidence scores for each class:")
            for idx, score in enumerate(predictions[0]):
                print(f"Class {idx} ({KNOWN_FACES['classes'][idx]}): {score:.4f}")
            
            return class_index, confidence

        except Exception as e:
            print(f"Error in predict_face: {str(e)}")
            return None, 0.0

    def recognize_face(self, image_data):
        print("\nStarting face recognition process...")
        frame = self.process_image(image_data)
        faces = self.detect_faces(frame)
        
        results = []
        if len(faces) == 0:
            print("No faces detected in the image")
            return [("No face detected", None, 0)], frame

        for (x, y, w, h) in faces:
            print(f"\nProcessing face at coordinates: x={x}, y={y}, w={w}, h={h}")
            face_crop = frame[y:y+h, x:x+w]
            
            # Save the cropped face for debugging (optional)
            debug_path = "debug_face.jpg"
            cv2.imwrite(debug_path, face_crop)
            print(f"Saved debug face image to {debug_path}")

            class_index, confidence = self.predict_face(face_crop)
            
            if class_index is not None and confidence > 0.5:  # Lowered threshold for testing
                name = KNOWN_FACES["classes"][class_index]
                print(f"Face recognized as {name} with confidence {confidence:.4f}")
                results.append((name, (x, y, w, h), confidence))
            else:
                print("Face not recognized or low confidence")
                results.append(("Unknown", (x, y, w, h), confidence))

        return results, frame

class AttendanceLog:
    def __init__(self, entry_log_file, exit_log_file):
        self.entry_log_file = entry_log_file
        self.exit_log_file = exit_log_file
        self._initialize_log_files()

    def _initialize_log_files(self):
        for log_file in [self.entry_log_file, self.exit_log_file]:
            if not os.path.exists(log_file):
                wb = Workbook()
                ws = wb.active
                ws.append(["Name", "Date", "Timestamp"])
                wb.save(log_file)

    def log_attendance(self, name, action):
        log_file = self.entry_log_file if action == "enter" else self.exit_log_file
        now = datetime.now()
        date_today = now.strftime("%Y-%m-%d")
        time_stamp = now.strftime("%Y-%m-%d %H:%M:%S")

        try:
            wb = load_workbook(log_file)
            ws = wb.active

            # Check for duplicate entries
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == name and row[1] == date_today:
                    print(f"Duplicate entry prevented for {name} on {date_today}")
                    return False

            ws.append([name, date_today, time_stamp])
            wb.save(log_file)
            print(f"Successfully logged attendance for {name} at {time_stamp}")
            return True
        except Exception as e:
            print(f"Error logging attendance: {e}")
            return False

class FaceRecognitionApp:
    def __init__(self, app, model_path, entry_log_file, exit_log_file):
        self.app = app
        self.face_recognition = FaceRecognition(model_path)
        self.attendance_log = AttendanceLog(entry_log_file, exit_log_file)
        self.setup_routes()

    def setup_routes(self):
        self.app.route("/")(self.index)
        self.app.route("/recognize_face", methods=["POST"])(self.recognize_face)

    def index(self):
        return render_template("index.html")

    def recognize_face(self):
        try:
            print("\nReceived recognition request")
            data = request.json
            image_data = data.get("image")
            action = data.get("action")
            
            if not image_data or not action:
                return jsonify({"error": "Invalid data"}), 400

            results, frame = self.face_recognition.recognize_face(image_data)
            print(f"Recognition results: {results}")

            if not results or results[0][0] == "No face detected":
                return jsonify({"message": "No face detected"}), 200

            response_data = []
            for name, (x, y, w, h), confidence in results:
                if name != "Unknown":
                    logged = self.attendance_log.log_attendance(name, action)
                    status = "logged" if logged else "already logged"
                    response_data.append({
                        "name": name,
                        "confidence": float(confidence),
                        "status": status
                    })
                
                # Draw bounding box and label
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 3)
                label = f"{name} ({confidence:.2f})"
                cv2.putText(frame, label, (x, y - 10), 
                          cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

            _, img_encoded = cv2.imencode('.jpg', frame)
            img_base64 = base64.b64encode(img_encoded).decode('utf-8')

            return jsonify({
                "message": f"Attendance processed for {', '.join([r['name'] for r in response_data])}",
                "results": response_data,
                "image": f"data:image/jpeg;base64,{img_base64}"
            }), 200

        except Exception as e:
            print(f"Error in /recognize_face: {str(e)}")
            return jsonify({"error": str(e)}), 500

# Initialize FaceRecognitionApp
face_recognition_app = FaceRecognitionApp(app, CNN_MODEL_PATH, ENTRY_LOG_FILE, EXIT_LOG_FILE)

if __name__ == "__main__":
    app.run(debug=True)